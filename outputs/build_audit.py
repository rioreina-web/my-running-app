import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

rows=[]
with open('_runs_raw.txt') as f:
    for line in f:
        line=line.rstrip('\n')
        if not line: continue
        p=line.split('|')
        rows.append(p)

# columns: date,type,source,miles,laps,fast_laps,pblk,strava,voice,display,bucket
wb=Workbook()

# ---- Summary sheet ----
s=wb.active; s.title='Summary'
ARIAL=lambda **k: Font(name='Arial', **k)
ink='262220'; coral='E8643C'; paper='F4F1EA'
hdrfill=PatternFill('solid', fgColor='262220')
s['A1']='Run display audit — every run, what the app actually shows'
s['A1'].font=ARIAL(bold=True, size=14, color=ink)
s['A2']=f'{len(rows)} runs · Jan 17 – Jun 20, 2026 · replicates the app’s own rep-detection logic'
s['A2'].font=ARIAL(size=10, color='6B6259')

# buckets summary with formulas referencing the data sheet
s['A4']='Outcome'; s['B4']='Runs'; s['C4']='% of all'
for c in ('A4','B4','C4'):
    s[c].font=ARIAL(bold=True, color='FFFFFF'); s[c].fill=hdrfill; s[c].alignment=Alignment(horizontal='left')
buckets=['Renders','Aerobic run, data but no chart','No data (voice/no GPS)']
labels={'Renders':'Shows a rep chart (works)',
        'Aerobic run, data but no chart':'EMPTY — has full GPS data, but no chart drawn (fixable display bug)',
        'No data (voice/no GPS)':'EMPTY — no split data exists (voice-only / pre-Strava)'}
r=5
for b in buckets:
    s[f'A{r}']=labels[b]
    s[f'B{r}']=f'=COUNTIF(Runs!K:K,"{b}")'
    s[f'C{r}']=f'=B{r}/$B$8'
    s[f'A{r}'].font=ARIAL(color=ink); s[f'B{r}'].font=ARIAL(color=ink); s[f'C{r}'].font=ARIAL(color=ink)
    s[f'C{r}'].number_format='0%'
    r+=1
s['A8']='Total'; s['B8']=f'=COUNTA(Runs!A2:A{len(rows)+1})'
s['A8'].font=ARIAL(bold=True,color=ink); s['B8'].font=ARIAL(bold=True,color=ink)
s['A10']='“EMPTY” total (rows you tap and see the one-sentence message):'
s['A10'].font=ARIAL(italic=True,color=coral)
s['B10']='=B6+B7'; s['B10'].font=ARIAL(bold=True,color=coral)

s.column_dimensions['A'].width=62; s.column_dimensions['B'].width=10; s.column_dimensions['C'].width=10

# ---- Runs sheet ----
d=wb.create_sheet('Runs')
headers=['Date','Type','Source','Miles','Laps','Fast laps (<6:10/mi)','Parsed blocks','Strava-linked','Voice note','What the sheet shows','Why']
d.append(headers)
for i,c in enumerate(headers,1):
    cell=d.cell(row=1,column=i); cell.font=ARIAL(bold=True,color='FFFFFF'); cell.fill=hdrfill
    cell.alignment=Alignment(horizontal='center',wrap_text=True,vertical='center')
greenf=PatternFill('solid',fgColor='E3EFE3'); amberf=PatternFill('solid',fgColor='FBEBD9'); greyf=PatternFill('solid',fgColor='ECE9E3')
for p in rows:
    date,typ,src,miles,laps,fast,pblk,strava,voice,disp,bucket=p
    d.append([date,typ,src,float(miles) if miles else None,int(laps),int(fast),int(pblk),
              'Yes' if strava=='Y' else 'No','Yes' if voice=='Y' else 'No',disp,bucket])
    rr=d.max_row
    fill=greenf if bucket=='Renders' else (amberf if bucket.startswith('Aerobic') else greyf)
    for col in range(1,12):
        cl=d.cell(row=rr,column=col); cl.font=ARIAL(size=10,color=ink); cl.fill=fill
    d.cell(row=rr,column=4).number_format='0.00'
    d.cell(row=rr,column=10).font=ARIAL(size=10,bold=(disp=='EMPTY'),color=(coral if disp=='EMPTY' else ink))
widths=[12,11,11,8,7,12,9,9,9,16,30]
for i,w in enumerate(widths,1):
    d.column_dimensions[chr(64+i)].width=w
d.freeze_panes='A2'
d.auto_filter.ref=f'A1:K{len(rows)+1}'

wb.save('run_display_audit.xlsx')
print('rows',len(rows))
